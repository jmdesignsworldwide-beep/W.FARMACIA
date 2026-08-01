#!/usr/bin/env python3
"""
Genera un archivo de inventario FEO A PROPÓSITO para probar el importador
(Tanda 3 · Pieza 3). Reproduce lo que Marien pidió y lo que Wilkins mandará:

  • filas de totales en el medio
  • encabezados repetidos
  • celdas combinadas (título y subtotales)
  • precios con símbolo de peso
  • números dominicanos ambiguos (1,250.50 y 1.250,50 en el mismo archivo)
  • fechas en dos formatos + una como fecha real de Excel (serial)
  • nombres con espacios de más
  • productos y lotes mezclados en la MISMA hoja
  • filas basura / vacías, y filas incompletas (sin costo, sin lote)

Uso:  pip install openpyxl  &&  python3 docs/fixtures/generar_fixture.py
Genera:  inventario-sucio-ejemplo.xlsx  y  inventario-sucio-ejemplo.csv
NO es dependencia del proyecto: openpyxl es solo para regenerar el fixture.
"""
import csv
import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font

AQUI = os.path.dirname(os.path.abspath(__file__))

# Encabezados sucios, tal como los escribe un humano (no limpios).
HEADER = ['Descripción', 'Principio', 'P. Venta', 'Laboratorio', 'Cant', 'Vence', 'Costo', 'Lote']

# Cada fila es lo que el importador tiene que sobrevivir. (valor crudo, tal cual)
# Nota: 'Vence' mezcla texto DD/MM/AAAA, texto ambiguo, guiones, y una fecha real.
FECHA_REAL = datetime.datetime(2027, 6, 30)  # openpyxl la escribe como serial de Excel
FILAS = [
    ['  Losartán 50 mg  ', 'Losartán', 'RD$ 51.00', 'Genfar', 120, '15/03/2027', '42.00', 'L-2201'],
    ['Amoxicilina 500mg', 'Amoxicilina', '1.250,50', 'MK', '1,000', FECHA_REAL, '980,00', 'AMX-99'],
    ['Ibuprofeno   400mg', 'Ibuprofeno', 'RD$ 1,250.50', 'Genfar', 60, '03/04/2027', '', ''],  # ambiguo mes/día, sin costo/lote
    ['', '', '', '', '', '', '', ''],  # fila vacía / basura
    ['TOTAL PASILLO 1', '', 'RD$ 2,552.00', '', '', '', '', ''],  # fila de TOTALES en el medio
    ['Descripción', 'Principio', 'P. Venta', 'Laboratorio', 'Cant', 'Vence', 'Costo', 'Lote'],  # encabezado REPETIDO
    ['Metformina 850 mg', 'Metformina', '95.00', 'Rowe', 200, '01-12-2026', '70.00', 'MET-01'],  # fecha con guiones
    ['Losartán 50 mg', 'Losartán', '', 'Genfar', 40, '20/09/2027', '42.00', 'L-2202'],  # MISMO producto, OTRO lote
    ['  Omeprazol 20 mg', 'Omeprazol', 'RD$ 180.00', 'Genven', '', '', '', ''],  # solo catálogo (sin cantidad/lote)
    ['Acetaminofén 500 mg', 'Acetaminofén', '1,250.50', 'MK', 500, '30/06/2027', '900.00', 'ACE-7'],  # decimal con punto
    ['Azitromicina 500 mg', 'Azitromicina', 'RD$ 320,00', 'Pfizer', 30, '05/05/2027', '', 'AZI-3'],  # peso + coma decimal, sin costo
    ['xxx', 'xxx', 'xxx', '', '', '', '', ''],  # basura
    ['TOTAL GENERAL', '', 'RD$ 5,000.00', '', '', '', '', ''],  # otra fila de totales
]


def escribir_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    # Título en celda COMBINADA (A1:H1)
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = 'FARMACIA WILKINS — INVENTARIO (hecho a mano)'
    c.font = Font(bold=True, size=14)
    ws.append([])  # fila 2 vacía
    ws.append(HEADER)  # fila 3: encabezado
    for f in FILAS:
        ws.append(f)
    # Combina la celda de la fila de totales del medio ("TOTAL PASILLO 1")
    # (queda a la altura correcta según lo escrito arriba)
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value in ('TOTAL PASILLO 1', 'TOTAL GENERAL'):
            r = row[0].row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    out = os.path.join(AQUI, 'inventario-sucio-ejemplo.xlsx')
    wb.save(out)
    print('escrito', out)


def escribir_csv():
    # CSV con la TRAMPA de configuración español: separador punto y coma.
    out = os.path.join(AQUI, 'inventario-sucio-ejemplo.csv')
    with open(out, 'w', newline='', encoding='utf-8') as fh:
        w = csv.writer(fh, delimiter=';')
        w.writerow(['FARMACIA WILKINS - INVENTARIO'])
        w.writerow([])
        w.writerow(HEADER)
        for f in FILAS:
            # La fecha real se escribe como texto DD/MM/AAAA en el CSV
            fila = [v.strftime('%d/%m/%Y') if isinstance(v, datetime.datetime) else v for v in f]
            w.writerow(fila)
    print('escrito', out)


if __name__ == '__main__':
    escribir_xlsx()
    escribir_csv()
